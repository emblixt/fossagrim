import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import numpy as np
import os
import unittest

from Fossagrim.utils.definitions import heureka_mandatory_standdata_keys, \
    heureka_standdata_keys, heureka_standdata_desc, \
    heureka_treatment_keys, heureka_treatment_desc, \
    fossagrim_standdata_keys, m3fub_to_m3sk, \
    translate_keys_from_fossagrim_to_heureka
from Fossagrim.utils.monetization_parameters import \
    parameters, variables_used_in_monetization, calculation_part1, resampled_section, \
    money_value, cbo_flow, project_benefits, buffer, fossagrim_values, forest_owner_values

import Fossagrim.plotting.misc_plots as fpp


def example_gis_database():
    """
    This is just a note on how to start handling tables from gis (exported from Allma)
    :return:
    """
    import geopandas as gpd
    # TODO

    base_dir = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\Sandbox\\Allma\\TEST Nord-Odal\\04180001900020000.gdb"

    # The following loop goes through all data tables in the above gis project and locates the table that contains
    # the key 'BEST_NR'
    for filename in os.listdir(base_dir):
        if "gdbtable" in filename:
            print('Reading {}'.format(filename))
            data = gpd.read_file(os.path.join(base_dir, filename))
            if 'BEST_NR' in list(data.keys()):
                break
    # And the following line finds the index of bestand nr. 279
    index = np.where(data['BEST_NR'] == 279.0)


def my_float(x):
    if isinstance(x, str):
        return float(x.replace(',', '.'))
    else:
        return float(x)


def my_str(x):
    if x is None:
        return ''
    else:
        return str(x)


def get_row_index(table, key, item):
    """
    Returns the row index of the item
    :param table:
        panda DataFrame
    :param key:
        str
        Name of key we are searching within
    :param item:
        str
        The name we are looking for

    :return:
    """
    for i, var in enumerate(table[key]):
        if item in var:
            return i

    return None


def read_excel(filename, header, sheet_name):
    try:
        table = pd.read_excel(filename, header=header, sheet_name=sheet_name, engine='openpyxl')
    except PermissionError as err_msg:
        print(err_msg)
        return None

    # Clean keys from trailing spaces or returns
    old_keys = list(table.keys())
    table = table.rename(columns={_key: _key.strip() for _key in old_keys})

    return table


def write_csv_file(write_to_file, default_keys, default_desc, append=False, **kwargs):
    """
    Writes the parameter values from the kwargs to the 'write_to_file'.
    It writes to a semicolon separated data file whose first line contains a description, and second
    line the key names, then each line is a line of data

    Example output is the StandData.csv files that is used by Heureka for import of stand
    data

    :param write_to_file:
        Name of csv file to write data to
    :param default_keys:
        list
        List of key names that the csv file must include, eg. heureka_standdata_keys
    :param default_desc:
        list
        List of descriptions that the csv file can include, eg. heureka_standdata_desc
    :param append:
        bool
        if True, the data in kwargs are appended to an existing output file without creating a header
    :param kwargs:
        list of keyword arguments.
        keywords must match the list 'heureka_standdata_keys' given in utils.definitions.py
        The value of each key is used when writing the 'write_to_file' file
    :return:
    """
    if len(default_keys) != len(default_desc):
        raise IOError('The lists of keys and descriptions must be of same length')

    # overwrite any existing file, and create the header.
    if not append:
        with open(write_to_file, 'w') as f:
            f.write(';'.join(default_desc) + '\n')
            f.write(';'.join(default_keys) + '\n')

    # arrange the data from kwargs
    data = [''] * len(default_keys)
    for key in kwargs:
        if key not in default_keys:
            print('WARNING: key "{}" not found in accepted default keys'.format(key))
            continue
        this_data = kwargs[key]
        if isinstance(this_data, float):
            this_string = '{:.2f}'.format(this_data)
        else:
            this_string = str(this_data)
        data[default_keys.index(key)] = this_string

    # append the data
    if len(kwargs) > 0:  # only write data if given
        with open(write_to_file, 'a') as f:
            f.write(';'.join(data) + '\n')


def read_raw_heureka_results(filename, sheet_name, read_only_these_variables=None, verbose=False):
    """
    Read heureka results as exported from Heureka (by simply copying a simulation result and pasting it
    into an empty sheet in Excel) and returns the data in a transposed DataFrame with one line for each period (5 years)
    NOTE, the Variable 'Treatments:Year' must be included in the raw results

    :param filename:
    :param sheet_name:
    :param read_only_these_variables:
       list
       If provided, read_raw_heureka_results will only read the variables (given by name) contained in this list
    :param verbose:
        bool
        creates qc plot(s)
    :return:
        panda DataFrame

    """
    table = read_excel(filename, 0, sheet_name)
    # Some treatments in Heureka break the default periods of 5 years into smaller sub-periods, we choose to remove
    # these from the returned dataframe
    try:
        year_row = get_row_index(table, 'Variable', 'Year')
    except KeyError as error:
        print('Variable Year not in file! ', error)
        return None
    five_years = [np.mod(this_year, 5) == 0 for this_year in table.iloc[year_row, 3:]]
    # print(table.iloc[year_row, 3:][five_years])
    data_dict = {}
    unit_dict = {}
    for variable in list(table['Variable']):
        if read_only_these_variables is not None:
            if variable not in read_only_these_variables:
                continue
        row_i = get_row_index(table, 'Variable', variable)
        data_dict[variable] = table.iloc[row_i, 3:][five_years]
        unit_dict[variable] = table.iloc[row_i, 2]

    if verbose:
        qc_plot_dir = os.path.join(os.path.split(filename)[0], 'QC_plots')
        fpp.plot_raw_data(data_dict, sheet_name, 'Raw Data', qc_plot_dir)

        # fig, ax = plt.subplots(figsize=(12, 12))
        # _x = data_dict['Year']
        # for key, _y in data_dict.items():
        #     if key in ['Year', 'Treatment', 'Unit']:
        #         continue
        #     ax.plot(_x, _y, label=key)
        # ax.set_title('Raw Heureka {}'.format(sheet_name))
        # ax.set_xlabel('Year')
        # ax.grid(True)
        # ax.legend()
        # fig.savefig(os.path.join(qc_plot_dir, 'raw_heureka_{}.png'.format(sheet_name)))

    result = pd.DataFrame(data=data_dict)
    result.attrs = unit_dict
    return result


def combine_raw_heureka_results(filename, sheet_name, function, variables, verbose=False):
    """
    Uses read_heureka_results() to load all variables from a raw Heureka result,
    then combines the variables according to the input function, eg:
    > variables = {'x': 'Total Carbon Stock (dead wood, soil, trees, stumps and roots)',
                   'y': 'Soil Carbon Stock',
                   'z': 'XXX', ...}
    > def func(x, y, z):
        return x - y + 10. * z - 19.
    > combine_raw_heureka_results(filename, sheet_name, func, variables)

    THUS, the variables dictionary must have the same number of elements as the input function has parameters
    :param filename:
    :param sheet_name:
    :param function:
        function
        E.G.
        > def my_func(x, y):
        >   return x - y
    :param variables
        dict
        Dictionary of variable where each key is a parameter in the input function (e.g. 'x') and the
        value is the name of variable (column) in the raw heureka results file (e.g. 'Soil Carbon Stock')
        E.G.
        {'x': 'Total Carbon Stock (dead wood, soil, trees, stumps and roots)',
         'y': 'Soil Carbon Stock'}
    :param verbose:
    :return:
    """
    data = read_raw_heureka_results(filename, sheet_name, verbose=verbose)
    _variables = {_key: data[variables[_key]] for _key in list(variables.keys())}

    return function(**_variables)


def rearrange_raw_heureka_results(filename, sheet_names, combine_sheets, monetization_file=None, verbose=False):
    """

    :param filename:
        str
        Name of file with Heureka results
    :param sheet_names:
        list
        List of strings with names of sheets that contains raw results from different Heureka simulations
    :param combine_sheets:
        dict
        { <NAME OF COMBINED RESULTS>: [<SHEET NAME FOR FOREST TYPE 1 IN HEUREKA RAW RESULTS>, <FRACTION TYPE 1>,
                                       <SHEET NAME FOR FOREST TYPE 2 IN HEUREKA RAW RESULTS>, <FRACTION TYPE 2>,
                                       ... ],
         <ANOTHER COMBINED RESULT>: [<SHEET NAME FOR FOREST TYPE 1 IN HEUREKA RAW RESULTS>, <FRACTION TYPE 1>,
                                     <SHEET NAME FOR FOREST TYPE 2 IN HEUREKA RAW RESULTS>, <FRACTION TYPE 2>,
                                     ... ]}
    :param monetization_file:
        str
        Full path name of monetization file which we try to generate if not existing
    :param verbose:
        bool
        creates qc plot(s)
    :return:
    """
    from openpyxl import load_workbook

    result_sheet_name = 'Rearranged results'

    rearrange_result_file = False

    # Test if the excel file with Heureka results already contain Rearranged results
    wb = load_workbook(filename)
    if result_sheet_name in wb.sheetnames:
        print("WARNING the sheet '{}' already exists in {}. Remove it before trying again.".format(
            result_sheet_name,
            os.path.basename(filename)))
        # TODO
        # The data in the old sheet is not removed before overwriting, so old data can hang on also after 'over writing'
        # This needs to be fixed before allowing for overwrite
        # _ans = input("Do you want to overwrite that sheet? Y/[N]:")
        # if _ans.upper() == 'Y':
        #     rearrange_result_file = True
    else:
        rearrange_result_file = True

    write_monetization_file = False
    if monetization_file is not None:
        if os.path.isfile(monetization_file):
            print("WARNING monetization file {} already exists.".format(
                os.path.split(monetization_file)[-1]))
            _ans = input("Do you want to overwrite? Y/[N]:")
            if _ans.upper() == 'Y':
                write_monetization_file = True
        else:
            write_monetization_file = True

        if write_monetization_file:
            # Create empty monetization file
            writer = pd.ExcelWriter(monetization_file, engine='xlsxwriter')
            wb = writer.book
            for _sheet in [result_sheet_name, 'Parameters', 'Monetization']:
                ws = wb.add_worksheet(_sheet)
            writer.close()

    start_cols = []
    for i, sheet_name in enumerate(sheet_names):
        table = read_raw_heureka_results(
            filename,
            sheet_name,
            read_only_these_variables=variables_used_in_monetization,
            verbose=verbose)
        this_start_col = i * (len(list(table.keys())) + 2)
        if rearrange_result_file:
            with pd.ExcelWriter(filename, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                table.to_excel(writer, sheet_name=result_sheet_name, startcol=this_start_col, startrow=2)
        start_cols.append(this_start_col)

    if rearrange_result_file:
        wb = load_workbook(filename)
        ws = wb[result_sheet_name]
        for i, header in enumerate(sheet_names):
            ws.cell(1, start_cols[i] + 1).value = header
        wb.save(filename)

    if combine_sheets is not None:
        last_start_col = start_cols[-1]
        # All the tables to be combined need to have the same (name and numbers) of columns
        for i, this_combined_result in enumerate(list(combine_sheets.keys())):
            fractions = []
            tables = []
            for j, sheet_name in enumerate(combine_sheets[this_combined_result]):
                if np.mod(j, 2) != 0:
                    # sheet names are found in positions 0, 2, 4, ... in the list
                    # On position 1, 3, 5, ... the fractions are stored
                    fractions.append(sheet_name)
                    continue
                tables.append(read_raw_heureka_results(filename, sheet_name, variables_used_in_monetization))
            combined_dict = {}
            for key in list(tables[0].keys()):
                if key == 'Treatment':
                    combined_dict[key] = tables[0][key].values
                    continue
                this_column = np.zeros(len(tables[0][key]))
                for k in range(len(tables)):
                    # Combine the tables using the fractions
                    this_column = this_column + tables[k][key].values * fractions[k]
                combined_dict[key] = this_column
            combined_table = pd.DataFrame(data=combined_dict)

            # Start writing the combined table
            this_start_col = last_start_col + (i + 1) * (len(list(combined_table.keys())) + 2)
            if rearrange_result_file:
                with pd.ExcelWriter(filename, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                    combined_table.to_excel(writer, sheet_name=result_sheet_name, startcol=this_start_col, startrow=2)

            if rearrange_result_file:
                wb = load_workbook(filename)
                ws = wb[result_sheet_name]
                ws.cell(1, this_start_col + 1).value = this_combined_result
                wb.save(filename)

            if write_monetization_file:
                _start_col = 0 + (i + 0) * (len(list(combined_table.keys())) + 2)
                with pd.ExcelWriter(monetization_file, mode='a', if_sheet_exists='overlay',
                                    engine='openpyxl') as writer:
                    combined_table.to_excel(writer, sheet_name=result_sheet_name, startcol=_start_col, startrow=2)

                wb = load_workbook(monetization_file)
                ws = wb[result_sheet_name]
                ws.cell(1, _start_col + 1).value = this_combined_result
                wb.save(monetization_file)

            if verbose:
                qc_plot_dir = os.path.join(os.path.split(filename)[0], 'QC_plots')
                fpp.plot_raw_data(combined_dict, this_combined_result, 'Combined', qc_plot_dir)

    return write_monetization_file


def load_heureka_results(filename, header, sheet_name=None, year_key=None, data_key=None):
    """
    Loads a typical output from Heureka where a time serie is given as a row.
    At line header + 1, there must be cells containing 'Variable' and 'Period ...' that we use to
    identify the time line of our selected data

    :param filename:
    :param header:
    :param sheet_name:
    :param year_key:
    :param data_key:
    :return:
    """
    if sheet_name is None:
        sheet_name = 'Sheet1'
    if year_key is None:
        year_key = 'Year'
    if data_key is None:
        data_key = 'Total Carbon Stock (dead wood, soil, trees, stumps and roots)'

    table = read_excel(filename, header, sheet_name)
    if table is None:
        return None
    table_keys = list(table.keys())
    variable_keys = list(table['Variable'])

    year_index = variable_keys.index(year_key)
    data_index = variable_keys.index(data_key)

    output_years = []
    output_data = []

    for key in table_keys:
        if 'Period' not in key:
            continue
        output_years.append(table[key][year_index])
        output_data.append(table[key][data_index])

    return np.array(output_years), np.array(output_data)


def arrange_export_of_one_stand(row_index, table, extra_keywords_list=None):
    """
    Utility function that collects, translates, and converts, the Fossagrim data into a format
    that can be used by heureka
    :param row_index:
        int
        row index (beginning at 0) of table which we want to export
    :param table:
        pandas DataFrame
        Table with data as read in 'export_fossagrim_stand_to_heureka'
        Each row contains one stand
    :param extra_keywords_list:
        list
        List of non-default keywords from heureka_standdata_keys that we want to add
    :return:
        dict
        Dictionary of keyword: value pairs that can be used by 'write_csv_file'
    """
    key_translation = translate_keys_from_fossagrim_to_heureka()
    keyword_arguments = {}

    if extra_keywords_list is None:
        extra_keywords_list = []

    # loop over all mandatory keywords that heureka needs as input
    for key in heureka_mandatory_standdata_keys + extra_keywords_list:
        keyword_arguments[key] = ''  # Give every mandatory key a default empty value at first
        if key in list(key_translation.keys()):
            # Use the translated keyword to extract data from the Fossagrim table. NO CONVERSION YET!
            keyword_arguments[key] = table[key_translation[key]][row_index]

    # Now do the conversion from Fossagrim data and units to Heureka compatible data and units
    for key, value in list(keyword_arguments.items()):
        if key in ['CountyCode']:
            keyword_arguments[key] = 13
        if key in ['InventoryYear', 'N', 'SoilMoistureCode', 'VegetationType', 'Peat']:
            try:
                keyword_arguments[key] = int(value)
            except ValueError as err_msg:
                print('For key {}: {}'.format(key, err_msg))
        if key == 'ProdArea':
            keyword_arguments[key] = value * 0.1  # from daa to ha
        if key == 'SiteIndexSpecies':
            if value == 'Gran':
                keyword_arguments[key] = 'G'
            elif value == 'Furu':
                keyword_arguments[key] = 'T'
            elif value == 'Bjørk':
                keyword_arguments[key] = 'B'
            else:
                raise NotImplementedError('Species "{}" is not recognized as either Gran or Furu'.format(value))
        if key == 'V':
            keyword_arguments[key] = value * 10. * m3fub_to_m3sk  # m3fub/daa to m3sk/ha
        if key == 'N':
            keyword_arguments[key] = value * 10.  # trees/mål to trees/ha
        if key == 'PlantDensity':
            keyword_arguments[key] = value * 10.  # plants/daa to plants/ha
        if key == 'PropPine':
            keyword_arguments[key] = table['Furu'][row_index] / table['Total'][row_index]
        if key == 'PropSpruce':
            keyword_arguments[key] = table['Gran'][row_index] / table['Total'][row_index]
        if key == 'PropBirch':
            keyword_arguments[key] = table['Lauv'][row_index] / table['Total'][row_index]
        if key == 'StandId':
            keyword_arguments[key] = table['Fossagrim ID'][row_index]

    return keyword_arguments


def average_over_stands(average_over, table, stand_id_key, average_name, verbose=False):
    """
    Calculates the area weighted averages

    :param average_over:
        dict
        dictionary where each item contains a list of strings/int which determines which stands to use in calculating
        average stands. The key is the name of each individual average. eg.
        {
            'Spruce': [67, 68, 70, 72, 75],
            'Pine': [32, 56, 58]
        }
    :param table:
        pandas DataFrame
        Table with data as read in 'export_fossagrim_stand_to_heureka'
        Each row contains one stand
    :param stand_id_key:
        str
        Key that is used to identify the stands used in 'average_over'
    :param average_name:
        str
        Given main name of the calculated average stands. The key in 'average_over' is then added, e.g.
        '<average_name>-Spruce'
    :return:
        panda DataFrame
        A copy of the original input table, but with one line of data for each average stand, which is the
        area weighted averages over the selected stands
    """
    from Fossagrim.utils.definitions import fossagrim_keys_to_average_over as keys_to_average_over
    from Fossagrim.utils.definitions import fossagrim_keys_to_sum_over as keys_to_sum_over
    from Fossagrim.utils.definitions import fossagrim_keys_to_most_of as keys_to_most_of

    # create a new empty DataFrame that should hold the average values
    avg_table = pd.DataFrame(columns=[_key.strip() for _key in list(table.keys())])

    for avg_group in average_over:
        if verbose:
            print('Averaging group {}, which contains stands {}'.format(avg_group, average_over[avg_group]))
        # extract the row indexes over which the average is calculated for this group
        average_ind = []
        for i, stand_id in enumerate(table[stand_id_key]):
            if stand_id in average_over[avg_group]:
                average_ind.append(i)

        if len(average_ind) == 0:
            print('WARNING, no match for average group {}'.format(avg_group))
            continue

        this_row_of_data = []
        # continue using the above indexes to calculate the average values
        total_area = np.sum(table['Prod.areal'][average_ind])
        # for key in fossagrim_standdata_keys:
        for key in list(table.keys()):
            if key.strip() in ['Fossagrim ID', 'Bestand']:
                this_data = '{}{}'.format(average_name, avg_group)
                if verbose:
                    print(' {}: {}'.format(key, this_data))
                this_row_of_data.append(this_data)
            elif key.strip() in keys_to_sum_over:
                this_data = np.sum(table[key][average_ind])
                if verbose:
                    print(' Sum over: {}: {}'.format(key, this_data))
                this_row_of_data.append(this_data)
            elif key.strip() in keys_to_average_over:
                this_data = np.sum(table['Prod.areal'][average_ind] * table[key][average_ind]) / total_area
                if verbose:
                    print(' Average over:{}: {}'.format(key, this_data))
                this_row_of_data.append(this_data)
            elif key.strip() in keys_to_most_of:
                # An area weighted 'most of' calculation
                this_data = table[key][average_ind].array
                area_weighted_average = np.sum(this_data * table['Prod.areal'][average_ind]) / total_area
                # index of original data closest to the average
                ind = np.argmin((this_data - area_weighted_average) ** 2)
                if verbose:
                    print(' Most of: {}: {}'.format(key, this_data[ind]))
                this_row_of_data.append(this_data[ind])
            else:
                # for other parameters, just copy the first selected stand
                this_data = table[key][average_ind[0]]
                if verbose:
                    print(' Copy first: {}: {}'.format(key, this_data))
                this_row_of_data.append(this_data)

        avg_table.loc[len(avg_table)] = this_row_of_data

        # for key in fossagrim_standdata_keys:
        #     if key in ['Fossagrim ID', 'Bestand']:
        #         avg_table[key] = [average_name]
        #     elif key in keys_to_sum_over:
        #         avg_table[key] = np.sum(table[key][average_ind])
        #     elif key in keys_to_average_over:
        #         avg_table[key] = np.sum(table['Prod.areal'][average_ind] * table[key][average_ind]) / total_area
        #     elif key in keys_to_most_of:
        #         # An area weighted 'most of' calculation
        #         this_data = table[key][average_ind]
        #         area_weighted_average = np.sum(this_data * table['Prod.areal'][average_ind]) / total_area
        #         # index of original data closest to the average
        #         ind = np.argmin((this_data - area_weighted_average)**2)
        #         avg_table[key] = this_data[ind]
        #     else:
        #         # for other parameters, just copy the first selected stand
        #         avg_table[key] = [table[key][average_ind[0]]]

    return avg_table


def export_fossagrim_stand_to_heureka(read_from_file, write_to_file, this_stand_only=None, average_over=None,
                                      stand_id_key=None,
                                      header=None, sheet_name=None, average_name=None, verbose=False):
    """
    Load a 'typical' Fossagrim stand data file (Bestandsutvalg) and writes an output file which Heureka can use
    See https://www.heurekaslu.se/wiki/Import_of_stand_register for description of parameters used by Heureka


    :param read_from_file:
    :param write_to_file:
    :param this_stand_only:
        str
        Name of stand to export
    :param average_over:
        dict
        dictionary where each item contains a list of strings/int which determines which stands to use in calculating
        average stands. The key is the name of each individual average. eg.
        {
            'Spruce': [67, 68, 70, 72, 75],
            'Pine': [32, 56, 58]
        }
    :param stand_id_key:
        str
        Key that is used to identify the stands used in 'average_over'
    :param header:
    :param sheet_name:
    :return:
    """
    if stand_id_key is None:
        stand_id_key = 'Bestand'
    if sheet_name is None:
        sheet_name = 0  # Reads only first sheet, opposite to pandas default were None reads all sheets.
    if header is None:
        header = 7

    table = read_excel(read_from_file, header, sheet_name)
    if table is None:
        print('WARNING, stands could not be loaded from {}'.format(read_from_file))
        return None

    append = False
    notes = [''] * len(table[stand_id_key])

    if (average_over is not None) and isinstance(average_over, dict):  # export averaged stands
        this_stand_only = None
        table = average_over_stands(average_over, table, stand_id_key, average_name, verbose=verbose)
        notes = [
            'Area weighted average for {} of stands {}'.format(_key, ', '.join([str(_x) for _x in average_over[_key]]))
            for _key in list(average_over.keys())]

    number_of_stands_written = 0
    for i, stand_id in enumerate(table[stand_id_key]):
        if table['Bonitering\ntreslag'][i] in ['Uproduktiv', '-']:  # or np.isnan(table['HovedNr'][i]):
            if verbose:
                print('Skipping stand {}'.format(stand_id))
            continue
        if (this_stand_only is not None) and (this_stand_only != stand_id):
            if verbose:
                print('Skipping stand {}'.format(stand_id))
            continue
        if verbose:
            print('Attempting to write {} to {}'.format(stand_id, write_to_file))
        # get the data from the fossagrim table arranged for writing in heureka format
        extra_keywords_list = ['PlantDensity', 'UserDefinedVariable2_RotationPeriod', 'UserDefinedVariable3_ThinningYear']
        keyword_arguments = arrange_export_of_one_stand(i, table, extra_keywords_list)
        write_csv_file(
            write_to_file,
            heureka_standdata_keys,
            heureka_standdata_desc,
            append=append,
            Note=notes[i],
            **keyword_arguments
        )
        number_of_stands_written += 1
        append = True

    if number_of_stands_written == 0:
        print('WARNING: No stands written to {}'.format(write_to_file))


def export_fossagrim_treatment(read_from_file, write_to_file, this_stand_only=None,
                               average_over=None, stand_id_key=None, header=None, sheet_name=None, average_name=None):
    """
    Load a 'typical' Fossagrim stand data file (Bestandsutvalg) and writes a 'treatment proposal' output file.
    This treatment proposal file does not fully follow the Heureka standard for TreatmentProposals but contains Fossagrim
    specific changes
    See https://www.heurekaslu.se/wiki/Import_of_stand_register for description of parameters used by Heureka
    and https://www.heurekaslu.se/help/index.htm?importera_atgardsforslag.htm for treatment proposals


    :param read_from_file:
    :param write_to_file:
    :param this_stand_only:
        str
        Name of stand to export
    :param average_over:
        dict
        dictionary where each item contains a list of strings/int which determines which stands to use in calculating
        average stands. The key is the name of each individual average. eg.
        {
            'Spruce': [67, 68, 70, 72, 75],
            'Pine': [32, 56, 58]
        }
    :param stand_id_key:
        str
        Key that is used to identify the stands used in 'average_over'
    :param header:
    :param sheet_name:
    :return:
    """
    if stand_id_key is None:
        stand_id_key = 'Bestand'
    if sheet_name is None:
        sheet_name = 0  # Reads only first sheet, opposite to pandas default were None reads all sheets.
    if header is None:
        header = 7
    if average_name is None:
        average_name = 'Avg Stand'

    table = read_excel(read_from_file, header, sheet_name)
    if table is None:
        return None

    # write to file, but only header lines
    write_csv_file(
        write_to_file,
        heureka_treatment_keys,
        heureka_treatment_desc,
        append=False
    )

    notes = [''] * len(table[stand_id_key])

    if (average_over is not None) and isinstance(average_over, dict):  # export averaged stands
        this_stand_only = None
        table = average_over_stands(average_over, table, stand_id_key, average_name)
        notes = [
            'Area weighted average for {} of stands {}'.format(_key, ', '.join([str(_x) for _x in average_over[_key]]))
            for _key in list(average_over.keys())]

    for i, stand_id in enumerate(table[stand_id_key]):
        if (table['Bonitering\ntreslag'][i] in ['Uproduktiv', '-']):  # or np.isnan(table['HovedNr'][i]):
            continue
        if (this_stand_only is not None) and (this_stand_only != stand_id):
            continue
        # Write the final felling at year 0 with subsequent planting
        write_csv_file(
            write_to_file, heureka_treatment_keys, heureka_treatment_desc, append=True,
            StandId=table['Fossagrim ID'][i], Year=0,
            Treatment='FinalFelling', Note=notes[i]
        )
        write_csv_file(
            write_to_file, heureka_treatment_keys, heureka_treatment_desc, append=True,
            StandId=table['Fossagrim ID'][i], Year=2,
            Treatment='Planting', PlantDensity=table['Plantetetthet'][i] * 10., Note=notes[i]
        )
        # write the thinning
        write_csv_file(
            write_to_file, heureka_treatment_keys, heureka_treatment_desc, append=True,
            StandId=table['Fossagrim ID'][i], Year=table['Tynnings år'][i],
            Treatment='Thinning', Note=notes[i]
        )
        # And the next final felling with subsequent planting
        write_csv_file(
            write_to_file, heureka_treatment_keys, heureka_treatment_desc, append=True,
            StandId=table['Fossagrim ID'][i], Year=table['Rotasjonsperiode'][i],
            Treatment='FinalFelling', Note=notes[i]
        )
        write_csv_file(
            write_to_file, heureka_treatment_keys, heureka_treatment_desc, append=True,
            StandId=table['Fossagrim ID'][i], Year=table['Rotasjonsperiode'][i] + 2,
            Treatment='Planting', PlantDensity=table['Plantetetthet'][i] * 10., Note=notes[i]
        )


def read_fossagrim_treatment(
        treatment_csv_file,
        project_tag,
        wood_species='Spruce'):
    """

    :param treatment_csv_file:
    :param project_tag:
    :param wood_species:
    :return:
        list
        [thinning_year, final_felling_year, plant_density]  # [year, year, plants/ha]
    """
    thinning_year = None
    final_felling_year = None
    plant_density = None
    with open(treatment_csv_file) as f:
        lines = f.readlines()
        for line in lines:
            split_line = line.split(';')
            if '{} {}'.format(project_tag, wood_species) == split_line[2]:
                if 'FinalFelling' in line:
                    final_felling_year = float(split_line[4])
                if 'Thinning' in line:
                    thinning_year = float(split_line[4])
                if 'Planting' in line:
                    plant_density = float(split_line[8])
    return [thinning_year, final_felling_year, plant_density]


def get_active_forests_from_stand(stand_file, stand_id_key=None, header=None, sheet_name=None):
    """
    Reads the stand file ("Bestandsutvalg") and finds out which forest stands are active and should be included,
    their (dominant) wood species, and their respective active area and the

    :param stand_file:
    :param stand_id_key:
    :param header:
    :param sheet_name:
    :return:
    """
    if stand_id_key is None:
        stand_id_key = 'Bestand'
    if sheet_name is None:
        sheet_name = 0  # Reads only first sheet, opposite to pandas default were None reads all sheets.
    if header is None:
        header = 7

    table = read_excel(stand_file, header, sheet_name)
    if table is None:
        print('WARNING, stands could not be loaded from {}'.format(stand_file))
        return None

    active_forests = {'Spruce': [], 'Pine': [], 'Birch': []}
    active_prod_areal = {'Spruce': [], 'Pine': [], 'Birch': []}  # Unit daa in Bestandsutvalg
    active_prod_areal_fractions = {}

    for i, stand_id in enumerate(table[stand_id_key]):
        this_wood_species = None
        if table['Bonitering\ntreslag'][i] in ['Gran', 'G', 'Spruce']:
            this_wood_species = 'Spruce'
        elif table['Bonitering\ntreslag'][i] in ['Furu', 'F', 'Pine']:
            this_wood_species = 'Pine'
        elif table['Bonitering\ntreslag'][i] in ['Bjørk', 'B', 'Birch']:
            this_wood_species = 'Birch'

        if table['Volum status'][i] == 1:
            active_forests[this_wood_species].append(stand_id)
            active_prod_areal[this_wood_species].append(my_float(table['Prod.areal'][i]))

    # Remove empty wood species from the active forests
    for key in list(active_forests.keys()):
        if len(active_forests[key]) == 0:
            _ = active_forests.pop(key)
            _ = active_prod_areal.pop(key)

    # Calculate total active area, and the fraction of each wood species against the total
    tot_area_per_species = {}
    tot_area = 0.
    for key in list(active_prod_areal.keys()):
        this_sum = sum(active_prod_areal[key])
        tot_area_per_species[key] = this_sum
        tot_area += this_sum
    for key in list(tot_area_per_species.keys()):
        active_prod_areal_fractions[key] = tot_area_per_species[key] / tot_area

    return active_forests, active_prod_areal, active_prod_areal_fractions


def get_kwargs_from_stand_OLD(stand_file, project_settings_file, project_tag):
    """
    Extracts necessary information from stand_file and project_settings_file to feed modify_monetization_file()
    :param stand_file:
        str
        Full path name of Fossagrim stand data file (Bestandsutvalg)
    :param project_settings_file:
        str
        Full path name of the project settings Excel sheet.
    :param project_tag:
        str
        Name tag that identifies the current project, e.g. "FHF23-007"
    :return:
        dict, list
        kwargs: Dictionary of kwargs needed by modify_monetization_file()
        coombine_fractions: list of combination fractions read from the stand_file
    """
    p_tabl = read_excel(project_settings_file, 1, 'Settings')
    i = None
    # Extract the project settings for the given project
    for i, p_name in enumerate(p_tabl['Project name']):
        if not isinstance(p_name, str):
            continue
        if project_tag in p_name:
            break

    combine_positions = [_x.strip() for _x in p_tabl['Position of fractions when combined'][i].split(',')]
    wb = openpyxl.load_workbook(stand_file, data_only=True)
    ws = wb.active
    combine_fractions = [ws[_x].value for _x in combine_positions]
    _kwarg_position_keys = [
        'Position of total area',
        'Position of Batch 1 total volume',
        'Position of Batch 2 total volume',
        'Position of passive forest total volume',
    ]
    _kwarg_direct_keys = [
        'Batch 1 start date',
        'Batch 2 delay',
        'Root net',
        'Contract length',
        'Rent',
        'Price growth',
        'Buffer',
        'Reserve years',
        'Net price',
        'Gross price'
    ]
    kwargs = {
        _key: ws[p_tabl[_key][i]].value for _key in _kwarg_position_keys
    }
    for _key in _kwarg_direct_keys:
        kwargs[_key] = p_tabl[_key][i]
    wb.close()

    import Fossagrim.io.fossagrim_io as fio
    return kwargs, combine_fractions


def get_kwargs_from_stand(stand_file, project_settings_file, project_tag):
    """
    Extracts necessary information from stand_file and project_settings_file to feed modify_monetization_file()
    :param stand_file:
        str
        Full path name of Fossagrim stand data file (Bestandsutvalg)
    :param project_settings_file:
        str
        Full path name of the project settings Excel sheet.
    :param project_tag:
        str
        Name tag that identifies the current project, e.g. "FHF23-007"
    :return:
        dict, list
        kwargs: Dictionary of kwargs needed by modify_monetization_file()
        combine_fractions: list of combination fractions read from the stand_file
    """
    p_tabl = read_excel(project_settings_file, 1, 'Settings')
    i = None
    # Extract the project settings for the given project
    for i, p_name in enumerate(p_tabl['Project name']):
        if not isinstance(p_name, str):
            continue
        if project_tag in p_name:
            break

    if i is None:
        raise IOError('Project tag {} not found in {}'.format(
            project_tag, os.path.basename(project_settings_file)
        ))

    combine_positions = [_x.strip() for _x in p_tabl['Position of fractions when combined'][i].split(',')]
    wb = openpyxl.load_workbook(stand_file, data_only=True)
    ws = wb.active
    combine_fractions = [ws[_x].value for _x in combine_positions]
    _kwarg_position_keys = [
        'Position of total area',
        'Position of Batch 1 total volume',
        'Position of Batch 2 total volume',
        'Position of passive forest total volume',
    ]
    _kwarg_direct_keys = {
        'Batch 1 start date': 'J',
        'Batch 2 delay': 'L',
        'Root net': 'N',
        'Contract length': 'O',
        'Rent': 'P',
        'Price growth': 'Q',
        'Buffer': 'R',
        'Reserve years': 'S',
        'Net price': 'T',
        'Gross price': 'U',
        'NIBOR 10yr': 'V'
    }

    try:
        kwargs = {
            _key: ws[p_tabl[_key][i]].value for _key in _kwarg_position_keys
        }
    except TypeError as error:
        print(error)
        print(i, _kwarg_position_keys)
        for _key in _kwarg_position_keys:
            print(_key, p_tabl[_key][i])
        raise TypeError(error)

    for _key in list(_kwarg_direct_keys.keys()):
        # kwargs[_key] = "='C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\[ProjectForestsSettings – WIP.xlsx]Settings'!${}${}".format(
        #     _kwarg_direct_keys[_key], i+3)  # +3 because 2 header lines and python starts counting at 0
        # Using the absolut path makes the linking works easier, but makes it difficult to share the file
        # UPDATE! If I open the linked file (ProjectForestSettings - WIP.xlsx) the linking works:-)
        kwargs[_key] = "='[{}]Settings'!${}${}".format(
            os.path.basename(project_settings_file),
            _kwarg_direct_keys[_key], i + 3)  # +3 because 2 header lines and python starts counting at 0
    wb.close()

    if None in combine_fractions:
        raise IOError('Some fractions are lacking in cells: {} in {}'.format(
            ', '.join(combine_positions), stand_file
        ))

    return kwargs, combine_fractions


def modify_monetization_file(write_to_file, **_kwargs):
    from openpyxl.workbook.defined_name import DefinedName

    with pd.ExcelWriter(write_to_file, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
        parameters.to_excel(
            writer, sheet_name='Parameters', startcol=0, startrow=1, index=False, header=False)
        calculation_part1.to_excel(
            writer, sheet_name='Monetization', startcol=0, startrow=0, index=False, header=False)
        resampled_section.to_excel(
            writer, sheet_name='Monetization', startcol=24, startrow=0, index=False, header=False)
        money_value.to_excel(
            writer, sheet_name='Monetization', startcol=30, startrow=1, index=False, header=False)
        cbo_flow.to_excel(
            writer, sheet_name='Monetization', startcol=35, startrow=1, index=False, header=False)
        project_benefits.to_excel(
            writer, sheet_name='Monetization', startcol=39, startrow=1, index=False, header=False)
        buffer.to_excel(
            writer, sheet_name='Monetization', startcol=49, startrow=2, index=False, header=False)
        fossagrim_values.to_excel(
            writer, sheet_name='Monetization', startcol=55, startrow=1, index=False, header=False)
        forest_owner_values.to_excel(
            writer, sheet_name='Monetization', startcol=66, startrow=1, index=False, header=False)

    # Modify monetization file with constants kwargs
    wb = openpyxl.load_workbook(write_to_file)

    ws = wb['Monetization']
    ws['F1'].value = _kwargs['Position of total area'] / 10.  # from mål to Ha
    ws['AF3'].value = _kwargs['Position of Batch 1 total volume']
    ws['AG3'].value = _kwargs['Position of Batch 2 total volume']
    ws['AI3'].value = _kwargs['Position of Batch 1 total volume'] + _kwargs['Position of Batch 2 total volume'] + \
                      _kwargs['Position of passive forest total volume']
    ws['AF5'].value = _kwargs['Batch 1 start date']
    ws['AG6'].value = _kwargs['Batch 2 delay']
    ws['AF8'].value = _kwargs['Root net']
    ws['AN4'].value = _kwargs['Contract length']
    ws['AO4'].value = _kwargs['Rent']
    ws['AP4'].value = _kwargs['Price growth']
    ws['AX4'].value = _kwargs['Buffer']
    ws['AY4'].value = _kwargs['Reserve years']
    ws['BH8'].value = _kwargs['Net price']
    ws['BI8'].value = _kwargs['Gross price']
    ws['BP3'].value = _kwargs['NIBOR 10yr']

    wb.save(write_to_file)

    style_monetization_file(write_to_file)

    # Create defined name (name manager) using openpyxl
    wb = openpyxl.load_workbook(write_to_file)
    hpap = openpyxl.workbook.defined_name.DefinedName("hpap", attr_text='Parameters!$B$2')
    wb.defined_names.add(hpap)
    hsawn = openpyxl.workbook.defined_name.DefinedName("hsawn", attr_text='Parameters!$B$3')
    wb.defined_names.add(hsawn)
    sfsawn = openpyxl.workbook.defined_name.DefinedName("SFsawn", attr_text='Parameters!$B$4')
    wb.defined_names.add(sfsawn)
    sfpp = openpyxl.workbook.defined_name.DefinedName("SFpp", attr_text='Parameters!$B$5')
    wb.defined_names.add(sfpp)
    sffuel = openpyxl.workbook.defined_name.DefinedName("SFfuel", attr_text='Parameters!$B$6')
    wb.defined_names.add(sffuel)
    p = openpyxl.workbook.defined_name.DefinedName("p", attr_text='Parameters!$B$7')
    wb.defined_names.add(p)
    pp = openpyxl.workbook.defined_name.DefinedName("pp", attr_text='Parameters!$B$8')
    wb.defined_names.add(pp)
    pf = openpyxl.workbook.defined_name.DefinedName("pf", attr_text='Parameters!$B$9')
    wb.defined_names.add(pf)
    psfuel = openpyxl.workbook.defined_name.DefinedName("psfuel", attr_text='Parameters!$B$10')
    wb.defined_names.add(psfuel)
    k = openpyxl.workbook.defined_name.DefinedName("k", attr_text='Parameters!$B$11')
    wb.defined_names.add(k)
    wb.save(write_to_file)


def style_monetization_file(write_to_file):
    from openpyxl.styles import NamedStyle
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from Fossagrim.utils.definitions import standard_colors as scrs

    colors = {
        scrs['time_1'].replace('#', ''): ['A1:G50'],
        scrs['products_1'].replace('#', ''): ['C3:F50'],
        scrs['substitution_1'].replace('#', ''): ['G3:G50'],
        scrs['bau_1'].replace('#', ''): ['H1:O50'],
        scrs['bau_2'].replace('#', ''): ['L2:O50'],
        scrs['project_case_1'].replace('#', ''): ['P1:Q50', 'BO2:BT110'],
        scrs['project_case_2'].replace('#', ''): ['Q2:Q50', 'BD2:BN110'],
        scrs['climate_benefit_1'].replace('#', ''): ['R1:W50', 'AN2:AW110'],
        scrs['climate_benefit_2'].replace('#', ''): ['U2:W50', 'AC1:AD110'],
        scrs['buffer_1'].replace('#', ''): ['Y1:AB110', 'AX2:BC110'],
        scrs['table_1'].replace('#', ''): ['AE2:AI9'],
        scrs['cbo_1'].replace('#', ''): ['AJ1:AM110']
    }

    # Create named styles
    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
    no_fill = openpyxl.styles.PatternFill(fill_type=None)
    no_side = openpyxl.styles.Side(border_style=None)
    side = openpyxl.styles.Side(border_style='thin')
    no_border = openpyxl.styles.borders.Border(left=no_side, right=no_side, top=no_side, bottom=no_side)
    all_border = openpyxl.styles.borders.Border(left=side, right=side, top=side, bottom=side)

    wb = openpyxl.load_workbook(write_to_file)

    # add the styles
    wb.add_named_style(date_style)

    ws = wb['Monetization']

    for _pos in ['AF5', 'AG5', 'AK4', 'AL4']:
        ws[_pos].style = 'date_style'
    for _pos in ['AF4', 'AG4', 'AH4', 'AI4', 'AK3', 'AL3', 'AM3', 'AO4', 'AP4', 'AQ4', 'AX4', 'BM2', 'BP3']:
        ws[_pos].number_format = '0.0%'
    for _row in np.arange(8, 109):
        ws['BK{}'.format(_row)].number_format = '0.0%'
        ws['BT{}'.format(_row)].number_format = '0.0%'
    for _row in np.arange(7, 109):
        ws['BQ{}'.format(_row)].number_format = 'kr #, ##0'
        ws['BR{}'.format(_row)].number_format = 'kr #, ##0'
        ws['BS{}'.format(_row)].number_format = 'kr #, ##0'

    for color in list(colors.keys()):
        for my_pos in colors[color]:
            for my_row in ws[my_pos]:
                for my_cell in my_row:
                    my_cell.fill = PatternFill(patternType='solid', fgColor=color)

    for my_cell_pos in ['A5', 'B5', 'F1', 'F2', 'AF3', 'AG3', 'AI3', 'AF5', 'AF8', 'AG6', 'AN4', 'AO4', 'AP4',
                        'AQ4', 'AT4', 'AX4', 'AY4', 'BH8', 'BI8', 'BP3']:
        ws[my_cell_pos].fill = no_fill
        ws[my_cell_pos].border = all_border

    for rng in ['C8:C50', 'P8:P50', 'H4:W5', 'BH9:BI50', 'BO6:BT50', 'BO2:BO2', 'BQ3:BR3']:
        for my_row in ws[rng]:
            for my_cell in my_row:
                my_cell.border = all_border

    for rng in ['BP7:BR7']:
        for my_row in ws[rng]:
            for my_cell in my_row:
                my_cell.font = Font(bold=True)

    for rng in ['H1:O1', 'P1:Q1', 'R1:W1', 'Y1:AB1', 'AC1:AD1',
                'H4:K4', 'L4:O4', 'BG4:BL4', 'BO2:BT2', 'BQ3:BQ5', 'BR3:BT5']:
        ws.merge_cells(rng)
        ws[rng.split(':')[0]].alignment = Alignment(horizontal='center', vertical='center')
        ws[rng.split(':')[0]].font = Font(bold=True)

    wb.save(write_to_file)


def qc_plots(monetization_file, project_tag, plot_dir=None):
    from Fossagrim.utils.definitions import standard_colors as scrs
    from Fossagrim.utils.definitions import standard_linestyles as sls
    from Fossagrim.utils.definitions import standard_linewidths as slw

    wb = openpyxl.load_workbook(monetization_file, data_only=True)
    ws = wb['Monetization']
    annual_climate_benefit_unit = ws['AC4'].value
    accum_climate_benefit_unit = ws['AD4'].value
    resampled_climate_benefit_unit = ws['AA4'].value
    overview_unit = ws['H4'].value
    contract_length = ws['AN4'].value
    wb.close()
    table = pd.read_excel(monetization_file, sheet_name='Monetization', header=5)
    if plot_dir is None:
        qc_plot_dir = os.path.join(os.path.split(monetization_file)[0], 'QC_plots')
    else:
        qc_plot_dir = plot_dir

    def annual_climate_benefit():
        x = table['t.1'].values
        y1 = table['30 yr contract'].values
        y2 = table['Unnamed: 28'].values
        y3 = table['100 yr contract'].values
        fig, ax = plt.subplots()
        ax.plot(x, y2, c=scrs['climate_benefit_0'],
                linestyle=sls['climate_benefit_0'],
                linewidth=slw['climate_benefit_0'],
                label='Climate benefit')
        ax.plot(x, y1, 'b-', label='30 yr contract')
        ax.plot(x, y3, 'b.', label='100 yr contract')
        ax.set_title('{} Annual Climate Benefit'.format(project_tag))
        ax.set_xlabel('Years')
        ax.set_ylabel(annual_climate_benefit_unit)
        ax.legend()
        ax.grid(True)
        fig.savefig(os.path.join(qc_plot_dir, '{} annual_climate_benefit.png'.format(project_tag)))

    def accumulated_climate_benefit():
        x = table['t.1'].values
        y1 = table['30 yr contract.1'].values
        y2 = table['Unnamed: 29'].values
        y3 = table['100 yr contract.1'].values
        fig, ax = plt.subplots()
        ax.plot(x, y2, c=scrs['climate_benefit_0'],
                linestyle=sls['climate_benefit_0'],
                linewidth=slw['climate_benefit_0'],
                label='Climate benefit')
        ax.plot(x, y1, 'b-', label='30 yr contract')
        ax.plot(x, y3, 'b.', label='100 yr contract')
        ax.set_title('{} Accumulated Climate Benefit'.format(project_tag))
        ax.set_xlabel('Years')
        ax.set_ylabel(accum_climate_benefit_unit)
        ax.legend()
        ax.grid(True)
        fig.savefig(os.path.join(qc_plot_dir, '{} accumulated_climate_benefit.png'.format(project_tag)))

    def resampled_climate_benefit():
        x = table['t.1'].values
        y1 = table['Unnamed: 27'].values
        y2 = table['Unnamed: 26'].values
        fig, ax = plt.subplots()
        ax.plot(x, y2, 'y.', label='Linear intpol / 5yr')
        ax.plot(x, y1, 'b-', label='Running average')
        ax.set_title('{} Resampled Climate Benefit'.format(project_tag))
        ax.set_xlabel('Years')
        ax.set_ylabel(resampled_climate_benefit_unit)
        ax.legend()
        ax.grid(True)
        fig.savefig(os.path.join(qc_plot_dir, '{} resampled_climate_benefit.png'.format(project_tag)))

    def overview():
        x = table['year'].values
        y1 = table['COPY OVER!.1'].values  # Base case
        y2 = table['COPY OVER!.2'].values  # Project case
        _y3 = table['Unnamed: 8'].values  # Product
        _y4 = table['Unnamed: 9'].values  # Substitution
        # shift the Product and Substitution arrays, and skip the last element
        y3 = np.insert(np.array([0.]), 1, _y3[:-1], axis=0)
        y4 = np.insert(np.array([0.]), 1, _y4[:-1], axis=0)
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(x, y1, c=scrs['bau_1'], label='Base case forest')
        ax.fill_between(x, y1, color=scrs['bau_1'])
        ax.plot(x, y1 + y3, c=scrs['products_1'], label='Base case product')
        ax.fill_between(x, y1 + y3, y1, color=scrs['products_1'])
        ax.plot(x, y1 + y4, c=scrs['substitution_0'], label='Base case substitution')
        ax.fill_between(x, y1 + y4, y1 + y3, color=scrs['substitution_0'])
        ax.plot(x, y2, c=scrs['project_case_0'], label='Project case')
        ax.set_title('{} Overview'.format(project_tag))
        ax.set_xlabel('Year')
        ax.set_ylabel(overview_unit)
        ax.set_xlim(np.nanmin(x) - 5, np.nanmin(x) + 155)
        ax.legend()
        ax.grid(True)
        fig.savefig(os.path.join(qc_plot_dir, '{} overview.png'.format(project_tag)))

    def farmed_offsets():
        x = table['t.1'].values
        y1 = table['Net farmed offsets / yr'].values
        y2 = table['Buffer reserved'].values
        y3 = table['Buffer released'].values
        y4 = table['Farmed offsets'].values
        fig, ax = plt.subplots(figsize=(10, 5))
        data = {
            'Net farmed offsets': y1,
            'Buffer reserved': y2,
            'Buffer released': y3
        }
        bottom = np.zeros(len(x))
        for label, height in data.items():
            ax.bar(x, height, label=label, bottom=bottom)
            bottom += height
        ax.plot(x, y4, label='Farmed offsets')
        ax.set_title('{} Farmed offsets'.format(project_tag))
        ax.set_ylabel('Ton')
        ax.set_xlabel('Time [years]')
        ax.legend()
        ax.grid(True)
        ax.set_xlim(0, contract_length)
        fig.savefig(os.path.join(qc_plot_dir, '{} farmed_offsets.png'.format(project_tag)))

    # annual_climate_benefit()
    # accumulated_climate_benefit()
    # resampled_climate_benefit()
    overview()
    # farmed_offsets()


def get_nature_and_climate_effect(result_file, stand_id, average_years=None, verbose=False):
    """
    Calculates the difference in predefined climate and nature parameters, between PRES and BAU,
    and integrates the result over the given number of years
    :param result_file:
         str
         Name of excel file that contain the "raw" Heureka results
    :param stand_id
         str
    :param average_years:
         int
         Number of years to calculate the integrated effect over
    """
    if average_years is None:
        average_years = 30

    # Heureka runs the simulation using timesteps which are 5 years long
    i = int(np.floor(average_years / 5))
    if verbose:
        print('average_years: {} corresponds to {} periods'.format(average_years, i))

    pres = None
    bau = None

    try:
        pres = read_raw_heureka_results(result_file, '{} {}'.format(stand_id, 'PRES'), verbose=verbose)
        if pres is None:
            raise ValueError('Missing Year variable')
    except ValueError as error:
        print(error)
        return None, None, None

    try:
        bau = read_raw_heureka_results(result_file, '{} {}'.format(stand_id, 'BAU'), verbose=verbose)
        if bau is None:
            raise ValueError('Missing Year variable')
    except ValueError as error:
        print(error)
        return None, None, None

    _c_diff = pres['Total Carbon Stock (dead wood, soil, trees, stumps and roots)'].values - pres['Soil Carbon Stock'].values - \
             bau['Total Carbon Stock (dead wood, soil, trees, stumps and roots)'].values + bau['Soil Carbon Stock'].values
    if verbose:
        print('Variables to sum over:')
        print('  Carbon (neglecting soil):', _c_diff[:i])
        # print('   Total carbon stock PRES:', pres['Total Carbon Stock (dead wood, soil, trees, stumps and roots)'][:i].values)
        # print('   Total carbon stock BAU:', bau['Total Carbon Stock (dead wood, soil, trees, stumps and roots)'][:i].values)
    c_diff = np.sum(_c_diff[:i])

    r_name = 'Recreation Index After'
    d_name1 = 'Dead Standing Trees >=20cm'
    d_name2 = 'Downed Deadwood >=20cm'

    if r_name in list(pres.keys()):
        _r_diff = pres[r_name].values - bau[r_name].values
        if verbose:
            print('  Recreation index:', _r_diff[:i])
            # print('   Recreation index PRES:', pres[r_name][:i].values)
            # print('   Recreation index BAU:',  bau[r_name][:i].values)
        r_diff = np.sum(_r_diff[:i])
    else:
        r_diff = None

    if d_name1 in list(pres.keys()):
        # _test = pres[d_name1].values + pres[d_name2].values - bau[d_name1].values - bau[d_name2].values
        _d_diff = pres[d_name1].values + pres[d_name2].values - bau[d_name1].values - bau[d_name2].values
        # _d_diff = pres[d_name1] + pres[d_name2] - bau[d_name1] - bau[d_name2]
        d_diff = np.sum(_d_diff[:i])
        if verbose:
            print('  Dead wood >= 20cm:', _d_diff[:i])
            # print('   TEST: ', _test[:i])
            # print('   Standing Dead wood PRES:', pres[d_name1][:i].values)
            # print('   Standing Dead wood BAU:', bau[d_name1][:i].values)
    else:
        d_diff = None

    return c_diff, d_diff, r_diff


def get_carbon_effect(result_file, stand_id,
                      average_years=None,
                      variable="Total Carbon Stock (dead wood, soil, trees, stumps and roots)",
                      ax=None,
                      save_plot_to=None):
    """
    Calculates the difference in Total carbon stock between PRES (preservation) and BAU (Business as usual)
    integrated over the given number of years
    :param result_file:
         str
         Name of excel file that contain the "raw" Heureka results
    :param stand_id
         str
    :param average_years:
         int
         Number of years to calculate the integrated effect over
    :param variable:
        str
        Variable name to use in the calculation of the carbon effect
    :param save_plot_to:
        str
        Name of file to save the plot to.
    :return:
        float
        Integrated effect of chosen parameter (variable) over <average_years> years for the given forest stand
        (project_tag) and wood_species
    """
    if average_years is None:
        average_years = 30
    try:
        diff = fpp.plot_from_heureka_results(
            result_file=result_file,
            sheets=['{} {}'.format(stand_id, _case) for _case in ['PRES', 'BAU']],
            params=variable,
            ax=ax,
            save_plot_to=save_plot_to,
            diff_sheets=True
        )
        # Heureka runs the simulation using timesteps which are 5 years long
        i = int(np.floor(average_years / 5))
        if diff is not None:
            return np.sum(diff[:i])
        else:
            return None
    except ValueError as error:
        print(error)
        return None


def collect_all_stand_data(
        out_file=None,
        base_dir=None,
        dry_run=True
):
    """
    Traverses the file structure for Heureka files and collects the stand data and selected Heureka
     results in one csv file

    out_file:
        str
        Full file name of output file
    :param base_dir:
        Path to directory where Heureka modelling results are stored (in sub folders of base_dir)
    :param dry_run:
        bool
        If True, not results are written, just checking for potential Heureka files
    :return:
    """
    # TODO Use this script to collect treatment too
    if base_dir is None:
        base_dir = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger"

    from pathlib import Path
    if dry_run and out_file is None:
        out_file = os.path.join(base_dir, 'TEST COLLECT STAND DATA.csv')

    if out_file is None:
        raise IOError('Name of output file must be given')

    # Creates a new Heureka compatible csv file for storing the data
    if not dry_run:
        write_csv_file(out_file, heureka_standdata_keys, heureka_standdata_desc, append=False)

    file_list = Path(base_dir).rglob('*Averaged stand data.csv')
    with open(out_file, 'a') as _out:  # append
        for stand_file in file_list:
            print('Working on {}'.format(stand_file.name))
            # see if there is a related Heureka results file
            heureka_result_file = str(stand_file).replace('Averaged stand data.csv', 'Heureka results.xlsx')
            if os.path.isfile(heureka_result_file):
                print(' Heureka result file exists')
            else:
                continue

            with open(stand_file, 'r') as _in:
                for line in _in.readlines():
                    split_line = line.split(';')
                    if 'FHF' in split_line[0]:
                        print('  Searching for stand id: {}'.format(split_line[0]))
                        carbon_effect, dead_wood_effect, recreation_effect = get_nature_and_climate_effect(
                            heureka_result_file,
                            split_line[0]
                        )
                        if carbon_effect is None:
                            alt_stand_id = split_line[0].replace('Avg Stand-', '')
                            print('  Searching for alternative stand id: {}'.format(alt_stand_id))
                            carbon_effect, dead_wood_effect, recreation_effect = get_nature_and_climate_effect(
                                heureka_result_file,
                                alt_stand_id
                            )
                        print('  ', carbon_effect)
                        split_line[101] = my_str(carbon_effect)
                        split_line[102] = my_str(dead_wood_effect)
                        split_line[103] = my_str(recreation_effect)
                        if not dry_run:
                            _out.write(';'.join(split_line))
                        else:
                            print('   - Dry run')


def collect_all_stand_data_OLD(
        out_file=None,
        base_dir=None,
        dry_run=True
):
    """
    Traverses the file structure for Heureka files and collects the stand data and selected Heureka
     results in one csv file

    out_file:
        str
        Full file name of output file
    :param base_dir:
        Path to directory where Heureka modelling results are stored (in sub folders of base_dir)
    :param dry_run:
        bool
        If True, not results are written, just checking for potential Heureka files
    :return:
    """
    if base_dir is None:
        base_dir = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger"

    if dry_run:
        # Create axes for plotting nothing
        fig, ax = plt.subplots()
    else:
        # get_carbon_effect() creates a new plot for each stand
        ax = None

    from pathlib import Path
    if dry_run and out_file is None:
        out_file = os.path.join(base_dir, 'TEST COLLECT STAND DATA.csv')

    if out_file is None:
        raise IOError('Name of output file must be given')

    file_list = Path(base_dir).rglob('*Averaged stand data.csv')
    with open(out_file, 'a') as _out:  # append
        for stand_file in file_list:
            print('Working on {}'.format(stand_file.name))
            # see if there is a related Heureka results file
            heureka_result_file = str(stand_file).replace('Averaged stand data.csv', 'Heureka results.xlsx')
            if os.path.isfile(heureka_result_file):
                print(' Heureka result file exists')
            else:
                continue

            with open(stand_file, 'r') as _in:
                for line in _in.readlines():
                    split_line = line.split(';')
                    if 'FHF' in split_line[0]:
                        print('  Searching for stand id: {}'.format(split_line[0]))
                        if dry_run:
                            save_plot = None
                        else:
                            save_plot = os.path.join(
                                os.path.dirname(heureka_result_file),
                                'Net Carbon Effect {}.png'.format(split_line[0]))

                        carbon_effect = get_carbon_effect(
                            heureka_result_file,
                            split_line[0],
                            ax=ax,
                            save_plot_to=save_plot
                        )
                        if carbon_effect is None:
                            alt_stand_id = split_line[0].replace('Avg Stand-', '')
                            print('  Searching for alternative stand id: {}'.format(alt_stand_id))
                            carbon_effect = get_carbon_effect(
                                heureka_result_file,
                                alt_stand_id,
                                ax=ax,
                                save_plot_to=save_plot
                            )

                        print('  ', carbon_effect)
                        split_line[101] = str(carbon_effect)
                        if not dry_run:
                            _out.write(';'.join(split_line))
                        else:
                            print('   - Dry run')


class TestCases(unittest.TestCase):
    def test_arrange_export(self):
        stand_file = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF24-0047 Åmli\\FHF24-0047 Forvaltningsplan_v02.xlsx"
        stand_id_key = 'Fossagrim ID'
        sheet_name='data'
        header = 7
        table = read_excel(stand_file, header, sheet_name)
        # print(list(table.keys()))
        keyword_args = (
            arrange_export_of_one_stand(1,
                                        table,
                                        ['PlantDensity', 'UserDefinedVariable2_RotationPeriod',
                                         'UserDefinedVariable3_ThinningYear']))
        print(keyword_args)

    def test_read_fossagrim_treatment(self):
        file = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF24-0014 Arne Tag\\FHF24-0014 Averaged treatment.csv"
        print(read_fossagrim_treatment(file, 'FHF24-0014', 'Spruce'))
        print(read_fossagrim_treatment(file, 'FHF24-0014', 'Rubbish'))

    def test_read_raw_heureka_results(self):
        file = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-999 Testing only\\FHF23-999 Heureka results.xlsx"
        sheet_name = 'FHF23-999 Pine BAU'
        result = read_raw_heureka_results(file, sheet_name, variables_used_in_monetization, False)
        print(result.keys())
        print(result)

    def test_modify_monetization_file(self):
        mf = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-999 Testing only\\Test Monetization.xlsx"
        sf = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-999 Testing only\\FHF23-999 Bestandsutvalg.xlsx"
        pf = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\ProjectForestsSettings - WIP.xlsx"
        tag = 'FHF23-999'

        # Create empty monetization file
        writer = pd.ExcelWriter(mf, engine='xlsxwriter')
        wb = writer.book
        writer.close()

        _kwargs, combine_fractions = get_kwargs_from_stand(sf, pf, tag)

        modify_monetization_file(mf, **_kwargs)

    def test_write_excel_with_equations(self):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.styles import Alignment

        colors = ['00660066', '00FFFFCC',
                  '00FF0000', '0000FF00', '00660066']
        fillers = []

        for color in colors:
            temp = PatternFill(patternType='solid',
                               fgColor=color)
            fillers.append(temp)

        dir_path = os.path.dirname(os.path.realpath(__file__))
        efile = os.path.join(dir_path, 'test.xlsx')
        sheet_names = ['One', 'Two', 'Three']

        writer = pd.ExcelWriter(efile, engine='xlsxwriter')
        wb = writer.book
        for sheet in sheet_names:
            ws = wb.add_worksheet(sheet)
        writer.close()

        wb = load_workbook(efile)
        ws = wb['Three']
        ws.cell(1, 1).value = 'Constant'
        ws.cell(1, 2).value = 4

        ws = wb['Two']
        ws.cell(1, 1).value = '=One!A1*One!B1'
        ws.cell(1, 2).value = '=One!A1*Three!$B$1'
        for i in range(4):
            ws.cell(i + 2, 1).value = '=One!A{}*One!B{}'.format(i + 2, i + 2)
            ws.cell(i + 2, 2).value = '=One!A{}*Three!$B$1'.format(i + 2)
            ws.cell(i + 2, 1).fill = fillers[0]
            ws.cell(i + 2, 2).fill = fillers[1]
        for my_row in ws['C6:D12']:
            for my_cell in my_row:
                my_cell.fill = fillers[3]

        ws.cell(1, 4).value = 'TEST'
        ws.cell(1, 4).alignment = Alignment(horizontal='center')
        ws.cell(1, 4).fill = fillers[2]
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8)
        wb.save(efile)

    def test_rearrange(self):
        file = "C:\\User\\marten\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-999 Testing only\\FHF23-999 Heureka results.xlsx"
        rearrange_raw_heureka_results(file, ['FHF23-999 Avg Stand-Pine BAU', 'FHF23-999 Avg Stand-Pine PRES'])

    def test_write_csv_file(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        write_csv_file(os.path.join(dir_path, 'test.csv'),
                       heureka_standdata_keys,
                       heureka_standdata_desc,
                       test=8, DGV=46, SiteIndexSpecies='G')
        write_csv_file(os.path.join(dir_path, 'test.csv'),
                       heureka_standdata_keys,
                       heureka_standdata_desc,
                       append=True, SiteIndexSpecies='T')
        write_csv_file(os.path.join(dir_path, 'test2.csv'),
                       heureka_standdata_keys,
                       heureka_standdata_desc)

    def test_export_fossagrim_stand(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        # read_from_file = "C:\\Users\\marten\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-003 Kloppmyra\\FHF23-003 Bestandsutvalg.xlsx"
        read_from_file = "C:\\Users\\marten\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-005 Kvistaul\\FHF23-005 Bestandsutvalg.xlsx"
        write_to_file = os.path.join(dir_path, 'standdata.csv')

        export_fossagrim_stand_to_heureka(read_from_file, write_to_file,
                                          stand_id_key='Fossagrim ID',
                                          average_over={
                                              'Spruce': ['FHF23-005-2', 'FHF23-005-13'],
                                              'Pine': ['FHF23-005-26', 'FHF23-005-27']
                                          }
                                          )
        # export_fossagrim_stand_to_heureka(read_from_file, write_to_file, average_over={'Spruce': [67, 70]},
        #                                   this_stand_only=67)

    def test_export_fossagrim_treatment(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        # read_from_file = "C:\\Users\\marten\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-003 Kloppmyra\\FHF23-003 Bestandsutvalg.xlsx"
        read_from_file = "C:\\Users\\marten\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-005 Kvistaul\\FHF23-005 Bestandsutvalg.xlsx"
        write_to_file = os.path.join(dir_path, 'treatment.csv')

        export_fossagrim_treatment(read_from_file, write_to_file,
                                   stand_id_key='Fossagrim ID',
                                   average_over={
                                       'Spruce': ['FHF23-005-2', 'FHF23-005-13'],
                                       'Pine': ['FHF23-005-26', 'FHF23-005-27']
                                   }
                                   )
        # export_fossagrim_treatment(read_from_file, write_to_file, average_over={'Spruce': [67, 70]},
        #                            this_stand_only=67)

    def test_get_kwargs_from_stand(self):
        kwargs, combine_fractions = get_kwargs_from_stand(
            # "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-999 Testing only\\FHF23-999 Bestandsutvalg.xlsx",
            # "C:\\Users\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF24-0014 Arne Tag\\Bestandsoversikt 16052024.xlsx",
            "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF24-0047 Åmli\\FHF24-0047 Forvaltningsplan_v02.xlsx",
            "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\ProjectForestsSettings.xlsx",
            "FHF24-0047 v02"
        )
        for _key in kwargs:
            print(_key, ':', kwargs[_key])
        print('-x-')
        print(combine_fractions)
        if None in combine_fractions:
            print('None is found')

        # f = 'C:\\tmp\\test.xlsx'
        # wb = openpyxl.load_workbook(f)
        # ws = wb['Sheet1']
        # for i, _key in enumerate(kwargs):
        #     ws['A{}'.format(i+1)] = _key
        #     ws['B{}'.format(i+1)] = kwargs[_key]
        # wb.save(f)

    def test_get_carbon_effect(self):
        # result_file = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF24-0016 Margrete Folsland\\FHF24-0016 Heureka results.xlsx"
        # project_tag = "FHF24-0016"
        # wood_species = "Pine"
        result_file = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF24-0047 Åmli\\FHF24-0047 v01 Heureka results.xlsx"
        project_tag = "FHF24-0047 v01 Pine"
        save_plot_to = result_file.replace("xlsx", "png")
        print(save_plot_to)
        # avg = get_carbon_effect(result_file, project_tag, wood_species=wood_species, save_plot_to=save_plot_to)
        avg = get_carbon_effect(result_file, project_tag, save_plot_to=save_plot_to)
        print(avg)
        plt.show()

    def test_combine_raw_herureka_results(self):
        file = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF23-999 Testing only\\FHF23-999 Heureka results.xlsx"
        sheet_name = 'FHF23-999 Pine BAU'

        def func(x, y):
            return x - y

        variables = {'x': 'Total Carbon Stock (dead wood, soil, trees, stumps and roots)',
                     'y': 'Soil Carbon Stock'}

        result = combine_raw_heureka_results(file, sheet_name, func, variables)
        print(result)

    def test_get_nature_carbon(self):
        file = "C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\FHF24-0047 Åmli\\FHF24-0047 v01 Heureka results.xlsx"
        stand_id = 'FHF24-0047 v01 Spruce'
        print(get_nature_and_climate_effect(file, stand_id, average_years=30, verbose=True))

    def test_collect_all_stand(self):
        collect_all_stand_data(out_file="C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\TEST COLLECT STAND DATA.csv",
                               dry_run=False)
