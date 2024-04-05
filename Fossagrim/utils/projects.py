import os
import numpy as np
import pandas as pd
import openpyxl
# from copy import deepcopy
import Fossagrim.io.fossagrim_io as fio

methods = ['BAU', 'PRES']  # "Business as usual" and "Preservation"


def arrange_import(_stand_file, _csv_stand_file, _csv_treatment_file, _average_over, _stand_id_key, _project_tag,
                   _verbose=False, **_kwargs):

    fio.export_fossagrim_stand_to_heureka(
        _stand_file,
        _csv_stand_file,
        average_over=_average_over,
        stand_id_key=_stand_id_key,
        average_name='{} Avg Stand'.format(_project_tag),
        verbose=_verbose
    )

    fio.export_fossagrim_treatment(
        _stand_file,
        _csv_treatment_file,
        average_over=_average_over,
        stand_id_key=_stand_id_key,
        average_name='{} Avg Stand'.format(_project_tag)
    )


def arrange_results(_result_file, _sheet_names, _combine_sheets, _monetization_file, _verbose=False, **_kwargs):
    # sheet names in the resulting Excel file
    write_monetization_file = \
        fio.rearrange_raw_heureka_results(
            _result_file, _sheet_names, _combine_sheets, monetization_file=_monetization_file, verbose=_verbose)

    if write_monetization_file:
        print('Try to modify Monetization file')
        fio.modify_monetization_file(_monetization_file, **_kwargs)


def project_settings(_project_tag, _project_settings_file):
    """
    Reads the project settings Excel file, and extracts information from it to be used in the Heureka simulation
    and sets up files necessary for storing results and monetization calculations for the given project

    Keys expected in project settings files:
    "Project name", "Status", "Forest types", "Average over stands", "Position of fractions when combined",
    "Check length of lists", "Stand id key", "Stands file", "Results file", "Monetization file"

    :param _project_tag:
        str
        Name tag that identifies the current project, e.g. "FHF23-007"

    :param _project_settings_file:
        str
        Full path name of the project settings Excel sheet.
    :return:
    """
    p_tabl = fio.read_excel(_project_settings_file, 1, 'Settings')
    w_dir = os.path.dirname(_project_settings_file)

    i = None
    for i, p_name in enumerate(p_tabl['Project name']):
        if not isinstance(p_name, str):
            continue
        if _project_tag in p_name:
            break

    _project_folder = os.path.join(w_dir, p_tabl['Project name'][i])
    _qc_folder = os.path.join(_project_folder, 'QC_plots')
    _stand_id_key = p_tabl['Stand id key'][i]
    _stand_file = os.path.join(_project_folder, p_tabl['Stands file'][i])
    _result_file = os.path.join(_project_folder, p_tabl['Results file'][i])
    _monetization_file = os.path.join(_project_folder, p_tabl['Monetization file'][i])
    _csv_stand_file = os.path.join(_project_folder, '{} Averaged stand data.csv'.format(_project_tag))
    _csv_treatment_file = os.path.join(_project_folder, '{} Averaged treatment.csv'.format(_project_tag))

    forest_types = [_x.strip() for _x in p_tabl['Forest types'][i].split(',')]
    average_over_strings = [_x.strip() for _x in p_tabl['Average over stands'][i].split(',')]
    if len(forest_types) != len(average_over_strings):
        raise ValueError('Number of forest types ({}) is not the same as Average stands ({})'.format(
            len(forest_types), len(average_over_strings)))
    _average_over = {}
    for j, f_type in enumerate(forest_types):
        if _stand_id_key == 'Bestand':
            _average_over[f_type] = [int(_x.strip()) for _x in average_over_strings[j].split(';')]
        else:
            _average_over[f_type] = [_x.strip() for _x in average_over_strings[j].split(';')]

    _result_sheets = \
        ['{} Avg Stand-{} {}'.format(_project_tag, _x, _y) for _x, _y in
         zip(np.repeat(list(_average_over.keys()), max(2, len(_average_over))), methods * len(_average_over))]

    # Create empty QC folder if it doesn't exist from before
    if not os.path.exists(_qc_folder):
        os.makedirs(_qc_folder)

    # Create empty results file
    create_results_file = False
    if os.path.isfile(_result_file):
        print("WARNING result file {} already exists.".format(
            os.path.split(_result_file)[-1]))
        _ans = input("Do you want to overwrite? Y/[N]:")
        if _ans.upper() == 'Y':
            create_results_file = True
    else:
        create_results_file = True

    if create_results_file:
        writer = pd.ExcelWriter(_result_file, engine='xlsxwriter')
        wb = writer.book
        for _sheet in _result_sheets:
            _ = wb.add_worksheet(_sheet)
        writer.close()

    # combine_positions = [_x.strip() for _x in p_tabl['Position of fractions when combined'][i].split(',')]
    # wb = openpyxl.load_workbook(_stand_file, data_only=True)
    # ws = wb.active
    # combine_fractions = [ws[_x].value for _x in combine_positions]
    # _kwarg_position_keys = [
    #     'Position of total area',
    #     'Position of Flow 1 total volume',
    #     'Position of Flow 2 total volume',
    #     'Position of passive forest total volume',
    # ]
    # _kwarg_direct_keys = [
    #     'Flow 1 start date',
    #     'Flow 2 delay',
    #     'Root net',
    #     'Contract length',
    #     'Rent',
    #     'Price growth',
    #     'Buffer',
    #     'Reserve years',
    #     'Net price',
    #     'Gross price'
    # ]
    # _kwargs = {
    #     _key: ws[p_tabl[_key][i]].value for _key in _kwarg_position_keys
    # }
    # for _key in _kwarg_direct_keys:
    #     _kwargs[_key] = p_tabl[_key][i]
    # wb.close()
    _kwargs, combine_fractions = fio.get_kwargs_from_stand(_stand_file, _project_settings_file, _project_tag)

    _combine_sheets = {}
    if len(combine_fractions) < 2:
        # only one forest type to "combine", so no combination of forest types necessary,
        # and the 'combine_fraction' is overridden using a factor of 1
        _combine_fractions = [1.]
    else:
        _combine_fractions = combine_fractions

    for j, method in enumerate(methods):
        _this_list = []
        for k, c_frac in enumerate(_combine_fractions):
            _this_list.append(_result_sheets[j + 2 * k])
            _this_list.append(c_frac)
        _combine_sheets['{} Combined Stands {} {}'.format(_project_tag, '-and-'.join(forest_types), method)] = \
            _this_list

    return _project_folder, _stand_file, _average_over, _stand_id_key, _result_file, _result_sheets, _combine_sheets, \
        _monetization_file, _csv_stand_file, _csv_treatment_file, _kwargs


if __name__ == '__main__':
    # project_tag = 'FHF23-999'
    project_tag = 'FHF24-006'
    project_settings_file = 'C:\\Users\\marte\\OneDrive - Fossagrim AS\\Prosjektskoger\\ProjectForestsSettings.xlsx'

    # Set to False after Heureka simulation results have been saved in result_file, and you want to
    # rearrange the results so that they are easier to include in Excel calculations
    fix_import = True

    verbose = True

    # For some reason, I need to open the Monetization file in Excel, and save it, before Python can
    # read it and create QC plots
    # So set this to True after opening and saving the Monetization file,
    monetization_file_has_been_opened_and_saved = True

    project_folder, stand_file, average_over, stand_id_key, result_file, result_sheets, combine_sheets, \
        monetization_file, csv_stand_file, csv_treatment_file, kwargs = \
        project_settings(project_tag, project_settings_file)

    if fix_import:
        arrange_import(stand_file, csv_stand_file, csv_treatment_file, average_over, stand_id_key, project_tag,
                       _verbose=verbose, **kwargs)
    elif (not fix_import) and (not monetization_file_has_been_opened_and_saved):
        arrange_results(result_file, result_sheets, combine_sheets, monetization_file, _verbose=verbose,  **kwargs)
    elif monetization_file_has_been_opened_and_saved:
        fio.qc_plots(monetization_file, project_tag)

